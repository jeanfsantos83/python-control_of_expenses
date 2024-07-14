import pandas as pd
from openpyxl import Workbook

class Despesa:
    def __init__(self, data, categoria, descricao, valor):
        self.data = data
        self.categoria = categoria
        self.descricao = descricao
        self.valor = valor

class ControleDespesas:
    def __init__(self):
        self.despesas = []

    def adicionar_despesa(self, despesa):
        self.despesas.append(despesa)

    def remover_despesa(self, data, categoria):
        for despesa in self.despesas:
            if despesa.data == data and despesa.categoria == categoria:
                self.despesas.remove(despesa)
                print(f"Despesa removida com sucesso!")
                return
        print(f"Despesa não encontrada.")

    def gerar_relatorio(self, arquivo_excel):
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Data'
        ws['B1'] = 'Categoria'
        ws['C1'] = 'Descrição'
        ws['D1'] = 'Valor'

        for i, despesa in enumerate(self.despesas, start=2):
            ws[f'A{i}'] = despesa.data
            ws[f'B{i}'] = despesa.categoria
            ws[f'C{i}'] = despesa.descricao
            ws[f'D{i}'] = despesa.valor

        wb.save(arquivo_excel)
        print(f"Relatório gerado com sucesso em {arquivo_excel}!")

def main():
    controle_despesas = ControleDespesas()

    while True:
        print("Menu:")
        print("  1. Adicionar despesa")
        print("  2. Remover despesa")
        print("  3. Gerar relatório")
        print("  4. Sair")
        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            data = input("Digite a data da despesa (DD/MM/YYYY): ")
            categoria = input("Digite a categoria da despesa: ")
            descricao = input("Digite a descrição da despesa: ")
            valor = float(input("Digite o valor da despesa: "))
            despesa = Despesa(data, categoria, descricao, valor)
            controle_despesas.adicionar_despesa(despesa)
            print("Despesa adicionada com sucesso!")
        elif opcao == "2":
            data = input("Digite a data da despesa para remover (DD/MM/YYYY): ")
            categoria = input("Digite a categoria da despesa para remover: ")
            controle_despesas.remover_despesa(data, categoria)
        elif opcao == "3":
            arquivo_excel = input("Digite o nome do arquivo Excel para gerar o relatório: ")
            controle_despesas.gerar_relatorio(arquivo_excel)
        elif opcao == "4":
            print("Saindo do programa...")
            break
        else:
            print("Opção inválida. Tente novamente.")

if __name__ == "__main__":
    main()
